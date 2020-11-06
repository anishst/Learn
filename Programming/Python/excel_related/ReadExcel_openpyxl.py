# https://automatetheboringstuff.com/chapter12/
#https://openpyxl.readthedocs.io/en/default/index.html

from openpyxl import load_workbook

wb = load_workbook(filename=r'Y:\Testing\Release 2.7\DataPurge\Data to be updated for Purge\Adjustment(WithDeposit)_Hist_DOI.xls')
# print(wb.get_sheet_names())
ws = wb.get_sheet_by_name("Tracing by Requirement")

#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Loop excel file: reads all but separate lines [ FINAL ]
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
for row in ws.iter_rows():
    for cell in row:
        print(cell.value, sep="\n")

#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# Loop excel file: need to specify number of columns [FINAL]
#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++
# for i in range(1,ws.max_row):
# 	print(i, ws.cell(row=i, column=1).value, ws.cell(row=i, column=2).value, sep="\t")


# Other options
# Maxrows = ws.max_row
# Maxcolumns = ws.max_column
# print(ws['A1'].value)
# print(ws.cell(row=1,column=2).value)

# Test versions


