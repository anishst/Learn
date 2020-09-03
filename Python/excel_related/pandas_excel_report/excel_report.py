# https://towardsdatascience.com/automate-excel-reporting-with-python-233dd61fb0f2

# Section 1 - Loading our Libraries
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference

# Section 2 - Loading our Data
df = pd.read_excel('https://github.com/datagy/pivot_table_pandas/raw/master/sample_pivot.xlsx', parse_dates=['Date'])
print(df.head())

# Section 3 - Testing Pivot Tables
filtered = df[df['Region'] == 'East']
quarterly_sales = pd.pivot_table(filtered, index = filtered['Date'].dt.quarter, columns = 'Type', values = 'Sales', aggfunc='sum')

print("Quarterly Sales Pivot Table:")
print(quarterly_sales.head())

# Section 04 - Creating and Excel Workbook
file_path = 'excel_report.xlsx'
quarterly_sales.to_excel(file_path, sheet_name = 'Quarterly Sales', startrow=3)

# Section 05 - Loading the Workbook
wb = load_workbook(file_path)
sheet1 = wb['Quarterly Sales']

# Section 06 - Formatting the First Sheet
sheet1['A1'] = 'Quarterly Sales'
sheet1['A2'] = 'datagy.io'
sheet1['A4'] = 'Quarter'

sheet1['A1'].style = 'Title'
sheet1['A2'].style = 'Headline 2'

for i in range(5, 9):
    sheet1[f'B{i}'].style='Currency'
    sheet1[f'C{i}'].style='Currency'
    sheet1[f'D{i}'].style='Currency'

# Section 07 - Adding a Bar Chart
bar_chart = BarChart()
data = Reference(sheet1, min_col=2, max_col=4, min_row=4, max_row=8)
categories = Reference(sheet1, min_col=1, max_col=1, min_row=5, max_row=8)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)
sheet1.add_chart(bar_chart, "F4")

bar_chart.title = 'Sales by Type'
bar_chart.style = 3
wb.save(filename = file_path)

# Section 08 - Getting Region Names
regions = list(df['Region'].unique())

# Section 09 - Looping Over All Regions
folder_path =  '.'

for region in regions:
    filtered = df[df['Region'] == f'{region}']
    quarterly_sales = pd.pivot_table(filtered, index=filtered['Date'].dt.quarter, columns='Type', values='Sales',
                                     aggfunc='sum')
    file_path = f"{folder_path}{region}.xlsx"
    quarterly_sales.to_excel(file_path, sheet_name='Quarterly Sales', startrow=3)

    wb = load_workbook(file_path)
    sheet1 = wb['Quarterly Sales']

    sheet1['A1'] = 'Quarterly Sales'
    sheet1['A2'] = 'datagy.io'
    sheet1['A4'] = 'Quarter'

    sheet1['A1'].style = 'Title'
    sheet1['A2'].style = 'Headline 2'

    for i in range(5, 10):
        sheet1[f'B{i}'].style = 'Currency'
        sheet1[f'C{i}'].style = 'Currency'
        sheet1[f'D{i}'].style = 'Currency'

    bar_chart = BarChart()
    data = Reference(sheet1, min_col=2, max_col=4, min_row=4, max_row=8)
    categories = Reference(sheet1, min_col=1, max_col=1, min_row=5, max_row=8)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    sheet1.add_chart(bar_chart, "F4")

    bar_chart.title = 'Sales by Type'
    bar_chart.style = 3
    wb.save(file_path)